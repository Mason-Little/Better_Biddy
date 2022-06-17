import sqlite3
import openpyxl

def Output_Running(Template):
    # Importing Workbook
    workbook = openpyxl.load_workbook(filename=Template[0])
    Ouput_Sheet = workbook.active

    # Connecting DB
    con = sqlite3.connect('mydatabase.db')
    cursorObj = con.cursor()

    # Fetching Data

    # Fetching Shots

    cursorObj.execute("select Shots_List from Input_Bids")
    results =cursorObj.fetchall()
    Shots_List = [str(result[0]) for result in results]

    # Fetching Production Notes
    cursorObj.execute('select Production_Notes_List from Input_Bids')
    results = cursorObj.fetchall()
    Production_Notes_List = [str(result[0]) for result in results]

    # Fetching Description Notes
    cursorObj.execute('select Description_List from Input_Bids')
    results = cursorObj.fetchall()
    Description_List = [str(result[0]) for result in results]

    # Finding Max Rows
    Input_Sheet_Max_Rows = Ouput_Sheet.max_row
    Input_Sheet_Max_Cols = Ouput_Sheet.max_column


    index = 0
    # Putting Back in Excel
    for i in range(17, 1000000, 4):
        if index == len(Shots_List) - 1:
            break
        Ouput_Sheet.cell(i, 2).value = Shots_List[index]
        index += 1
    index = 0
    for i in range(17, 1000000, 4):
        if index == len(Production_Notes_List) - 1:
            break
        Ouput_Sheet.cell(i, 5).value = Production_Notes_List[index]
        index += 1
    index = 0
    for i in range(17, 1000000, 4):
        if index == len(Description_List) - 1:
            break
        Ouput_Sheet.cell(i, 3).value = Description_List[index]
        index += 1

    cursorObj.execute("DELETE FROM Input_Bids")
    con.commit()

    workbook.save(filename = "C:\\Users\\masonl90\\Desktop\\Output.xlsx")


