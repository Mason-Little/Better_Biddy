import openpyxl
import sqlite3

def Input_Running(input):
    # Importing Workbook
    workbook = openpyxl.load_workbook(filename=input[0])
    Input_Sheet = workbook.active

    # Creating Data Base
    con = sqlite3.connect('mydatabase.db')
    cursorObj = con.cursor()
    cursorObj.execute("CREATE TABLE IF NOT EXISTS Input_Bids (Shots_List, Production_Notes_List, Description_List)")

    # Keywords for Bids
    shot_key_words = ["shot #", "shot id", "vfx #", 'shot name']
    Production_notes_key_words = ['production notes', 'description - vfx']
    Description_key_words = ['description', 'description - script', "vfx notes"]

    # Temp Lists
    Shots_List = []
    Production_Notes_List = []
    Description_List = []

    # Finding Max Rows
    Input_Sheet_Max_Rows = Input_Sheet.max_row
    Input_Sheet_Max_Cols = Input_Sheet.max_column


    # Finding Shots Numbers
    for i in range(1, Input_Sheet_Max_Cols + 1):
        variable = str(Input_Sheet.cell(1, i).value)
        if variable.lower() in shot_key_words:
            temp = str(Input_Sheet.cell(1, i))
            shots_column = ord(temp[15].lower()) - 96

    for i in range(1 + 1, Input_Sheet_Max_Rows + 1):
        if Input_Sheet.cell(i, shots_column).value is not None:
            Shots_List.append(Input_Sheet.cell(i, shots_column).value)

    # Finding Production Notes

    for i in range(2, Input_Sheet_Max_Rows + 1):
        variable = str(Input_Sheet.cell(1, i).value)
        if variable.lower() in Production_notes_key_words:
            temp = str(Input_Sheet.cell(1, i))
            shots_column = ord(temp[15].lower()) - 96

    for i in range(2, Input_Sheet_Max_Rows + 1):
        if Input_Sheet.cell(i, shots_column).value is not None:
            Production_Notes_List.append(Input_Sheet.cell(i, shots_column).value)

    # Finding Description
    for i in range(1, Input_Sheet_Max_Cols + 1):
        variable = str(Input_Sheet.cell(1, i).value)
        if variable.lower() in Description_key_words:
            temp = str(Input_Sheet.cell(1, i))
            shots_column = ord(temp[15].lower()) - 96

    for i in range(1 + 1, Input_Sheet_Max_Rows + 1):
        if Input_Sheet.cell(i, shots_column).value is not None:
            Description_List.append(Input_Sheet.cell(i, shots_column).value)


    # Inserting in DB
    for i in range(0, len(Shots_List)):
        cursorObj.execute("INSERT INTO Input_Bids VALUES(?, ?, ?)", (str(Shots_List[i]), str(Production_Notes_List[i]), str(Description_List[i])))
    con.commit()
